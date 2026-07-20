#!/usr/bin/env python3
"""Read case numbers from stdin and emit their official DOL visa classes as JSON."""

import json
import sys
from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("usage: extract-visa-class.py <official-disclosure.xlsx>")
    wanted = set(json.load(sys.stdin))
    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    case_index = header.index("CASE_NUMBER")
    visa_index = header.index("VISA_CLASS")
    found = {}
    for row in rows:
        case_number = row[case_index]
        if case_number in wanted:
            found[case_number] = row[visa_index]
            if len(found) == len(wanted):
                break
    json.dump(found, sys.stdout, sort_keys=True)


if __name__ == "__main__":
    main()
